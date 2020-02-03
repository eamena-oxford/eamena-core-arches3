from __future__ import division
import os
import csv
import json
import datetime
import string
import uuid
from django.conf import settings
from optparse import make_option
from django.core.management.base import BaseCommand, CommandError 
from openpyxl import load_workbook
import arches.app.models.models as archesmodels
from arches.management.commands import utils
from arches.app.utils.data_management.resources.formats.archesfile import ArchesReader
from django.contrib.gis.geos import GEOSGeometry
from arches.app.models.entity import Entity
from arches.app.models.models import UniqueIds
from django.core.exceptions import ObjectDoesNotExist

RELATION_HEADER = ['RESOURCEID_FROM', 'RESOURCEID_TO', 'START_DATE', 'END_DATE', 'RELATION_TYPE', 'NOTES']

class Command(BaseCommand):
    
    option_list = BaseCommand.option_list + (
        make_option('-o', '--operation', action='store', dest='operation', default='site_dataset',type='choice', choices=['validate'],help="'validate' must be chosen as the operation, and a 'validation_type' must be added as well"),
        make_option('-s', '--source', action='store', dest='source', default='',help='Directory containing the XLSX file you need to convert to .arches'),
        make_option('-d', '--dest_dir', action='store', dest='dest_dir',help='Directory, comprinsing of file name, where you want to save the .arches file'),
        make_option('-r', '--res_type', action='store', dest='resource_type', default='HERITAGE_RESOURCE_GROUP.E27', help='What kind of resource the source file contains e.g. HERITAGE_RESOURCE_GROUP.E27'),
        make_option('-a', '--append', action='store_true', dest='append_data', default=False, help='This flag indicates that the XLSX file contains data to append'),
        make_option('--validation_type',help='choose which type of validation to run')
    )
    
    def handle(self, *args, **options):
        
        print 'operation: '+ options['operation']
        if options['operation'] == 'validate':
            print '  suboperation: '+options['validation_type']
        package_name = settings.PACKAGE_NAME
        print 'package: '+ package_name
        start = datetime.datetime.now()

        workbook = load_workbook(options['source'], data_only=True)

        filelist = []
        if options['operation'] == 'validate':
            try:
                vtype = options['validation_type']
                fname = os.path.splitext(os.path.basename(options['dest_dir']))[0]
                error_path = os.path.join(settings.BULK_UPLOAD_DIR,
                    "{}_validation_errors-{}.json".format(fname,vtype)
                )

                if vtype == 'headers':
                    print "resourcetype"
                    print options['res_type']
                    result = self.validate_headers(workbook, 
                        skip_resourceid_col=options['append_data'],
                        resource_type=options['res_type'])
                    
                elif vtype == 'rows_and_values':
                    result = self.validate_rows_and_values(workbook)
                    
                elif vtype == 'dates':
                    data = self.get_values_for_validation(workbook,datatype="dates")
                    values = self.flatten_values(data)
                    result = self.validatedates(values)

                elif vtype == 'geometries':
                    data = self.get_values_for_validation(workbook,datatype="geomtries")
                    values = self.flatten_values(data)
                    result = self.validate_geometries(values)
                    
                elif vtype == 'concepts':
                    data = self.get_values_for_validation(workbook,datatype="domains")
                    result = self.validate_concepts(data)
                    
                elif vtype == 'files':
                    data = self.get_values_for_validation(workbook,datatype="files")
                    values = self.flatten_values(data)
                    result, filelist = self.validate_files(values)

                elif vtype == 'uniqueids':
                    data = self.get_values_for_validation(workbook, datatype='uniqueids')
                    values = self.flatten_values(data)
                    result = self.validate_uniqueids(values, options['res_type'])
                    
                elif vtype == 'write_arches_file':
                    result = self.write_arches_file(
                        workbook,options['res_type'],
                        options['dest_dir'],
                        options['append_data']
                    )

            except Exception as e:
                print repr(e)
            with open(error_path,'w') as out:
                json.dump(result,out)

        if filelist:
            self.stdout.write("Has attachments")
        
        print "elapsed time:",datetime.datetime.now()-start
        return
        
    def validatedates(self,date_tuples):
        '''wraps the existing date validation, only returns validation errors,
        no transformed date objects'''
        
        result = {'success':True,'errors':[]}
        for dt in date_tuples:
            if dt[0] == "x":
                continue
            validated = self.validatedate(dt[0])
            if not self.validatedate(dt[0]):
                msg = "{}: {} ({} > row {}, col {})".format(
                    dt[2],dt[0],dt[1],dt[3],dt[4]
                )
                result['errors'].append(msg)
        if result['errors']:
            result['success'] = False
        return result

    def validatedate(self, date, header = None, row_no = None):
        """validates a date string that has been passed from the spreadsheet.
        note that 'x' values should never be sent though this method. returns
        date objects."""
        
        valid_formats = [
            '%Y-%m-%d', #Checks for format  YYYY-MM-DD
            '%Y-%m-%d %X', #Checks for format  YYYY-MM-DD hh:mm:ss (locale)
            '%d-%m-%Y', #Checks for format  DD-MM-YYYY
            '%d/%m/%Y', #Checks for format  DD/MM/YYYY
            '%d/%m/%y', #Checks for format  DD/MM/YY
        ]
        
        ## check all valid full date formats
        for d_format in valid_formats:
            try:
                d = datetime.datetime.strptime(date, d_format)
                result = d.date()
                break
            except ValueError:
                result = None

        ## final attempt for just year format
        if not result:
            try:
                d = datetime.datetime.strptime(date,'%Y') #Checks for format  YYYY
                isodate = d.isoformat()
                d = isodate.strip().split("T")[0]
                result = d.date()
            except:
                result = None

        return result

    def validate_rows_and_values(self,workbook):
        '''Validates that the number of semicolon-separated values is consistent
        across a worksheet and spots empty cells'''
        result = {'success':True,'errors':[]}
        sheet_count = len(workbook.worksheets)
        rows_count  = 0
        ret = []
        offset_max = False

        for sheet_index,sheet in enumerate(workbook.worksheets):
            sheet_name = workbook.sheetnames[sheet_index]

            ## iterate from the last row to the first, to properly ignore
            ## empty rows that are on the end of the sheet.
            backward_rows = list(sheet.rows)
            backward_rows.reverse()
            for n,row in enumerate(backward_rows):
                if any([cell.value for cell in row]):
                    num_rows = len(backward_rows)-n
                    break
            rows_count = rows_count + num_rows

            #In the NOT tab, cells in a row are not part of the same merge
            # group, so the number of pipe-separated values need not be equal
            if not sheet_name == "NOT":
                ret = self.validate_value_number(sheet,sheet_name)
                for msg in ret:
                   result['errors'].append(msg)

            ## get the number of real columns based on the headers
            headers = [i.value for i in list(sheet.rows)[0] if i.value]

            ## now iterate the row normal direction to find empty cells
            for n,row in enumerate(list(sheet.rows)):
                if n+1 == num_rows:
                    break
                for cell in row[:len(headers)]:
                    msg = "Blank value in: {} row {}".format(sheet_name,n+1)
                    try:
                        if cell.value is None or str(cell.value).rstrip() == "":
                            result['errors'].append(msg)
                    except UnicodeEncodeError:
                        if cell.value.encode('ascii', 'ignore').rstrip() == "":
                            result['errors'].append(msg)

        if (rows_count/sheet_count).is_integer() is not True:
            result['errors'].append("Inconsistent number of rows across "\
            "workbook sheets.")

        if result['errors']:
            result['success'] = False
        return result
            
    def validate_value_number(self, sheet, sheet_name):

        msgs=[]
        num_cols = 0
        headers = []
        for row_index, row in enumerate(sheet.iter_rows()):
            if row_index == 0:
                headers = [cell.value for cell in row if cell.value]
                num_cols = len(headers)
                continue

            values_no = []

            # skip completely empty rows
            if not any(cell.value for cell in row):
                continue

            for col,cell in enumerate(row):
                # don't analyze values that are in columns without headers
                # (this is to control for situations where excel adds extra
                # empty columns somehow)
                if sheet_name == 'RELATIONS' and headers[col] == 'RESOURCEID_FROM':
                    continue

                if col+1 > num_cols:
                    continue
                if cell.value is not None:
                    value_encoded = (unicode(cell.value)).encode('utf-8')
                    cell_no = len(value_encoded.split("|"))
                    values_no.append(cell_no)
                else:
                    values_no.append(0)
            if values_no.count(values_no[0]) != len(values_no) or 0 in values_no:
                msgs.append("Inconsistent number of pipe-separated values:"\
                " {} row {}".format(sheet_name,row_index+1))

        return msgs

    def validate_headers(self, workbook, skip_resourceid_col = False, resource_type=None):

        if resource_type != 'relations':
            q = Entity().get_mapping_schema(resource_type)
        
        result = {'success':True,'errors':[]}
        for sheet in workbook.worksheets:
            if sheet.title == 'RELATIONS':
                restypenodes = RELATION_HEADER
            else:
                restypenodes = set(q.keys())
            for header in sheet.iter_cols(max_row = 1):
                nodename = header[0].value
                if nodename is not None:
                    if skip_resourceid_col == True and header[0].value =='RESOURCEID':
                        continue
                    if not nodename in restypenodes:
                        msg = "{} is not a valid {} node name".format(
                            nodename, resource_type)
                        result['errors'].append(msg)
        if result['errors']:
            result['success'] = False
        return result
        
    
    def validate_concepts(self,concept_data):
    
        result = {'success':True,'errors':[]}
        for sheet_name,contents in concept_data.iteritems():

            for node_name, concept_tuples in contents.iteritems():

                if sheet_name == 'RELATIONS' and node_name == 'RELATION_TYPE':
                    node_name = 'ARCHES_RESOURCE_CROSS-REFERENCE_RELATIONSHIP_TYPES.E55'

                label_lookup = self.get_label_lookup(node_name)
                for ct in concept_tuples:
                    if ct[0] == "x":
                        continue

                    #Values could be in Arabic or contain unicode chars, so it
                    #is essential to encode them properly.
                    value_encoded = (unicode(ct[0])).encode('utf-8')
                    for concept in value_encoded.split('|'):
                        concept = concept.rstrip().lstrip()
                        try:
                            label_lookup[concept.lower()]
                        except KeyError:
                            msg = "{} - {}: {} (row {}, col {})".format(
                                sheet_name,node_name,concept,ct[1],ct[2]
                            )
                            result['errors'].append(msg)

        if result['errors']:
            result['success'] = False
        return result

    def collect_concepts(self, node_conceptid, full_concept_list = []):
        ''' Collects a full list of child concepts given the conceptid of the node. Returns a list of a set of concepts, i.e. expounding the duplicates'''
        concepts_in_node = archesmodels.ConceptRelations.objects.filter(conceptidfrom = node_conceptid)
        # print len(concepts_in_node)
        if concepts_in_node.count() > 0:
            full_concept_list.append(node_conceptid) 
            for concept_in_node in concepts_in_node:
                
                self.collect_concepts(concept_in_node.conceptidto_id, full_concept_list)
        else:
            full_concept_list.append(node_conceptid)
        return list(set(full_concept_list)) 

    def validate_geometries(self,geom_tuples):
        '''wraps the existing geom validation, returns errors'''
        
        result = {'success':True,'errors':[]}
        for gt in geom_tuples:
            if gt[0] == "x":
                continue
            msg = self.validate_geometry(gt[0])
            if not validation == "valid":
                msg = "{}: {} ({} > row {}, col {})\n{}".format(
                    gt[2],gt[0],gt[1],gt[3],gt[4],msg
                )
                result['errors'].append(msg)

        if result['errors']:
            result['success'] = False
        return result
        
    def validate_geometry(self, geometry,header,row):
        """validates a geom string that has been passed from the spreadsheet.
        note that 'x' values should never be sent though this method. returns
        True or False."""
        try:
            GEOSGeometry(geometry)
            if GEOSGeometry(geometry).valid:
                return "valid"
            else:
                return "Intersecting polygon"
        except Exception as e:
            return str(e)

    def validate_files(self, file_tuples):
        """
        Going to validate files here. check they don't contain directory info and are an accepted file format.
        """
        result = {'success':True,'errors':[]}
        filelist = []
        for ft in file_tuples:
            filelist.append(ft[0])
            name, ext = os.path.splitext(os.path.basename(ft[0]))
            accepted_extensions = ['.pdf', '.jpg', '.png', '.doc', '.docx', '.txt']
            if ft[0] != name + ext:
                result['errors'].append("The file at header {}, line {} looks like it contains a folder name.".format(ft[2],ft[3]))
            if not ext in accepted_extensions:
                result['errors'].append("Cannot recognise file extension at header {}, line {}. Currently only accept {}".format(ft[2],ft[3], accepted_extensions))
        
        if result['errors']:
            result['success'] = False
        return result, filelist

    def validate_uniqueids(self, data, resource_type):
        """
        Going to validate that the unique ids are truly unique
        """
        result = {'success': True, 'errors': []}
        for row in data:
            eamena_id = row[0]
            if resource_type in settings.EAMENA_RESOURCES:
                id_type = settings.EAMENA_RESOURCES[resource_type]
            else:
                id_type = resource_type.split('_')[0]
            num = int(eamena_id[len(id_type) + 1:])
            found = UniqueIds.objects.filter(id_type=id_type, val=num)
            if len(found) > 0:
                result['errors'].append("The unique ID at header {}, line {} already exists within the database.".format(row[2], row[3]))

        if result['errors']:
            result['success'] = False
        return result

    def create_resourceid_list(self, workbook):
        ''''Looks for RESOURCEID column and creates a list of resourceids and returns '''
        resourceids_list = []
        for sheet_index,sheet in enumerate(workbook.worksheets):
            if workbook.sheetnames[sheet_index] == 'NOT':
                for col_index,header in enumerate(sheet.iter_cols(max_row = 1)):
                    if header[0].value =='RESOURCEID':
                        for row_index, row in enumerate(sheet.iter_rows(row_offset = 1)):
                            if row[col_index].value is not None:
                                resourceids_list.append(row[col_index].value)

        return resourceids_list
        
    def get_node_x_business_table_dict(self,wb):
        '''returns a dictionary that can be used to get the businesstablename
        of a node'''
        
        node_names = []
        for sheet in wb.worksheets:
            for header in sheet.iter_cols(max_row = 1):
                node_name = header[0].value
                if not node_name or node_name == 'RESOURCEID':
                    continue
                node_names.append(node_name)
        lookup = {}
        for name in node_names:
            try:
                obj = archesmodels.EntityTypes.objects.get(pk=name)
                lookup[name] = obj.businesstablename
            except ObjectDoesNotExist:
                if name in RELATION_HEADER:
                    if 'DATE' in name:
                        lookup[name] = 'dates'
                    elif name == 'RELATION_TYPE':
                        lookup[name] = 'domains'
                    else:
                        lookup[name] = 'strings'
        return lookup
        
    def flatten_values(self,data):
        '''flattens the dictionary output of get_values() into a list of tuples
        in the following format:
        
        (value,sheet_name,node_name,row_number,col_number)
        
        meant for easy iteration when the validation operation need not be grouped
        by node.
        '''
        
        output = []
        for sheet_name,contents in data.iteritems():
            for node_name,values in contents.iteritems():
                for val in values:
                    line = (val[0],sheet_name,node_name,val[1],val[2])
                    output.append(line)
        return output

    def get_values_for_validation(self,wb,datatype=''):
        '''collects all data values from the workbook, only of certain type
        if specified.
        '''
        
        col_letters1 = [i for i in string.ascii_uppercase]
        col_letters2 = [i*2 for i in string.ascii_uppercase]
        col_letters = col_letters1+col_letters2

        result = {}
        businesstable = self.get_node_x_business_table_dict(wb)

        for sheet_index,sheet in enumerate(wb.worksheets):
            sheet_name = wb.sheetnames[sheet_index]
            result[sheet_name] = {}
            for col_index,header in enumerate(sheet.iter_cols()):
                node_name = header[0].value

                if not node_name or node_name == 'RESOURCEID':
                    continue

                if datatype and businesstable[node_name] != datatype:
                    continue
                
                if not node_name in result[sheet_name]:
                    result[sheet_name][node_name] = []

                for row_index, row in enumerate(sheet.iter_rows(row_offset = 1)):
                    values = row[col_index].value
                    if not values:
                        continue
                    encoded_values = unicode(values).encode('utf-8')
                    for value in encoded_values.split("|"):
                        tuple = (value,row_index+2,col_letters[col_index])
                        result[sheet_name][node_name].append(tuple)

        return result
        
    def get_label_lookup(self, node_name, return_entity=False):
        
        node_obj = archesmodels.EntityTypes.objects.get(pk=node_name)
        all_concepts = self.collect_concepts(node_obj.conceptid_id,full_concept_list=[])

        ## dictionary will hold {label:concept.legacyoid} or {label:valueid}
        label_lookup = {}
        for c in all_concepts:
            cobj = archesmodels.Concepts.objects.get(pk=c)
            labels = archesmodels.Values.objects.filter(conceptid_id=c,valuetype_id="prefLabel")
            for label in labels:
                if return_entity:
                    label_lookup[label.value.lower()] = label.valueid
                else:
                    label_lookup[label.value.lower()] = cobj.legacyoid
                
        return label_lookup

    def get_existing_resource(self, eamena_id):
        '''From EAMENA_ID return the resource entityid'''

        try:
            uuid.UUID(eamena_id)
        except ValueError:
            id_type = eamena_id[:-settings.ID_LENGTH-1]
            num = int(eamena_id.split('-')[-1])
            uniqueid = archesmodels.UniqueIds.objects.get(id_type=id_type, val=num)
            parent = archesmodels.Relations.objects.get(entityidrange=uniqueid.entityid_id)
            return parent.entityiddomain.entityid
        else:
            archesmodels.Entities.objects.get(entityid=eamena_id, entitytypeid__isresource=True)
            return eamena_id

    def write_arches_file(self,workbook,resourcetype,destination,append=False):
        '''trimmed down version of SiteDataset, removing all validation operations.
        only produces a .arches file now.'''
        
        result = {'success': True, 'errors': []}
        ResourceList = []
        Relations = {}

        if append:
            resourceids_list = self.create_resourceid_list(workbook)

        for sheet_index,sheet in enumerate(workbook.worksheets):
            sheet_name = workbook.sheetnames[sheet_index]
            for col_index,header in enumerate(sheet.iter_cols(max_row = 1)):
                GroupNo = 0
                node_name = header[0].value
                if not node_name or node_name == 'RESOURCEID':
                    continue

                if sheet_name == 'RELATIONS' and node_name == 'RELATION_TYPE':
                    node_name = 'ARCHES_RESOURCE_CROSS-REFERENCE_RELATIONSHIP_TYPES.E55'

                try:
                    node_obj = archesmodels.EntityTypes.objects.get(pk=node_name)
                except ObjectDoesNotExist:
                    if 'DATE' in node_name:
                        entitytype = 'DATE'
                        datatype = 'dates'
                    else:
                        entitytype = 'TEXT'
                        datatype = 'strings'
                else:
                    entitytype = node_obj.entitytypeid
                    datatype = node_obj.businesstablename

                    label_lookup = self.get_label_lookup(node_name, return_entity=(resourcetype == 'relations' and sheet_name == 'RELATIONS'))

                for row_index, row in enumerate(sheet.iter_rows(row_offset = 1)):

                    value = row[col_index].value
                    if not value:
                        continue

                    resourceid = row_index if append == False else resourceids_list[row_index]

                    if resourceid not in Relations:
                        Relations[resourceid] = {
                            'RESOURCEID_FROM': [],
                            'RESOURCEID_TO': [],
                            'START_DATE': [],
                            'END_DATE': [],
                            'RELATION_TYPE': [],
                            'NOTES': []
                        }

                    encoded = unicode(value).encode('utf-8')
                    for concept in encoded.split('|'):
                        
                        try:
                            ## remove leading and trailing spaces
                            concept = concept.rstrip().lstrip()
                            GroupNo += 1

                            ## skip x in every case. this is a reserved placeholder to signify a non-value
                            if concept == "x":
                                continue

                            if sheet_name == "NOT":
                                GroupName = col_index
                            else:
                                GroupName = " ".join((sheet_name, str(GroupNo)))

                            ## data transformations must happen with domains and dates
                            if datatype == 'domains':
                                outval = label_lookup[concept.lower()]

                            elif datatype == 'dates':
                                outval = self.validatedate(concept)

                            ## anticipate django's auto renaming of saved files
                            elif "FILE_PATH" in entitytype:
                                outval = concept.replace(" ","_")

                            elif sheet_name == 'RELATIONS' and 'RESOURCEID' in node_name and '-' in concept:
                                outval = self.get_existing_resource(concept)

                            else:
                                outval = concept

                            if sheet_name == 'RELATIONS':
                                Relations[resourceid][header[0].value].append(outval)
                            else:
                                row = [str(resourceid),resourcetype,entitytype,outval, GroupName]
                                ResourceList.append(row)
                        
                        except Exception as e:
                            result['success'] = False
                            result['errors'].append(
                            "error writing value: {} - {}".format(
                                concept,entitytype)
                            )

        ## make relations file
        relationsfile = destination.replace(".arches", ".relations")
        with open(relationsfile, 'wb') as rel:
            if resourcetype == 'relations':
                writer = csv.writer(rel, delimiter=',')
            else:
                writer = csv.writer(rel, delimiter="|")
            writer.writerow(RELATION_HEADER)
            for resid in Relations:
                for i, val in enumerate(Relations[resid]['RESOURCEID_TO']):
                    reltype = Relations[resid]['RELATION_TYPE'][i]
                    fromid = resid
                    toid = val
                    if len(Relations[resid]['RESOURCEID_FROM']) > 0:
                        fromid = Relations[resid]['RESOURCEID_FROM'][0]
                    if '-' not in toid:
                        toid = int(val) - 2  # remove one for header line and one for index number
                    try:
                        startdate = Relations[resid]['START_DATE'][i]
                        enddate = Relations[resid]['END_DATE'][i]
                    except IndexError:
                        startdate = None
                        enddate = None
                    try:
                        notes = Relations[resid]['NOTES'][i]
                    except IndexError:
                        notes = ''
                    line = [fromid, toid, startdate, enddate, reltype, notes]
                    writer.writerow(line)

        if resourcetype != 'relations':
            if len(ResourceList) == 0:
                result['success'] = False
                result['errors'].append("No '%s' resources found to upload." % resourcetype)

            with open(destination, 'wb') as archesfile:
                writer = csv.writer(archesfile, delimiter ="|")
                writer.writerow(['RESOURCEID', 'RESOURCETYPE', 'ATTRIBUTENAME', 'ATTRIBUTEVALUE', 'GROUPID'])
                ResourceList = sorted(ResourceList, key = lambda row:(row[0],row[4],row[2]), reverse = False)
                # this code added to rearrange the featuregroups as a bandaid fix for featuregroups not displaying correctly in the edit forms.
                def reaarange_feature_group(group):
                    FFTRow = ''
                    FFTCRow = ''
                    FSTRow = ''
                    for i in group:
                        if 'FEATURE_FORM_TYPE.I4' in i:
                            FFTRow = group[group.index(i)]
                        if 'FEATURE_SHAPE_TYPE.E55' in i:
                            FSTRow = group[group.index(i)]
                        if 'FEATURE_FORM_TYPE_CERTAINTY.I6' in i:
                            FFTCRow = group[group.index(i)]
                    if FFTRow != '' and FSTRow != '' and FFTCRow != '':
                        group.remove(FFTRow)
                        group.remove(FSTRow)
                        group.remove(FFTCRow)
                        group.append(FFTRow)
                        group.append(FSTRow)
                        group.append(FFTCRow)
                    elif FFTRow != '' and FFTCRow != '' and FSTRow == '':
                        group.remove(FFTRow)
                        group.remove(FFTCRow)
                        FSTRow = [group[0][0], group[0][1], 'FEATURE_SHAPE_TYPE.E55', 'FEATURE_EVIDENCE_SHAPE_TYPE:1', group[0][4]]
                        group.append(FFTRow)
                        group.append(FSTRow)
                        group.append(FFTCRow)
                    elif FFTRow != '' and FSTRow != '' and FFTCRow == '':
                        group.remove(FFTRow)
                        group.remove(FSTRow)
                        group.append(FFTRow)
                        group.append(FSTRow)
                    elif FFTCRow != '' and FSTRow != '' and FFTRow == '':
                        group.remove(FSTRow)
                        group.remove(FFTCRow)
                        group.append(FFTCRow)
                        group.append(FSTRow)
                    else:
                        print 'This error is from ~ln 694 of commands/management/Excelreader.py'
                    return group
                from operator import itemgetter
                from itertools import groupby
                newResourceList = []
                feature_group_attributes = 'FEATURE_ARRANGEMENT_TYPE.E55','FEATURE_ASSIGNMENT_INVESTIGATOR_NAME.E41','FEATURE_FORM_TYPE.I4', 'FEATURE_FORM_TYPE_CERTAINTY.I6', 'FEATURE_NUMBER_TYPE.E55','FEATURE_SHAPE_TYPE.E55'
                for a, b in groupby(ResourceList, itemgetter(0)):
                    for c, d in groupby(b, itemgetter(4)):
                        group = list(d)
                        feature_group = False
                        for row in group:
                            if any(item in feature_group_attributes for item in row) == True:
                                feature_group = True
                        if feature_group == True:
                            group = reaarange_feature_group(group)
                        for r in group:
                            newResourceList.append(r)
                for row in newResourceList:
                    writer.writerow(row)

            arches_errors = ArchesReader().validate_file(destination,break_on_error=False)
            for ae in arches_errors:
                result['success'] = False
                result['errors'].append(".arches file validation error: "+ae)

        return result

